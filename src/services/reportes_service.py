from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
from src.db.repositories import productos_repository, movimientos_repository


def dividir_fecha_hora(valor):
    if not valor:
        return ("", "")
    try:
        dt = datetime.fromisoformat(valor.replace("Z", ""))
        return (dt.date().isoformat(), dt.time().strftime("%H:%M:%S"))
    except:
        return (valor, "")


def aplicar_estilos_excel(ws, num_columnas):
    """Aplicar estilos consistentes a todos los Excel"""
    color_header = "1a959f"
    color_fila_alterna = "e8f4f6"
    
    thin_border = Border(
        left=Side(style="thin", color="D3D3D3"),
        right=Side(style="thin", color="D3D3D3"),
        top=Side(style="thin", color="D3D3D3"),
        bottom=Side(style="thin", color="D3D3D3")
    )
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color=color_header, end_color=color_header, fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=num_columnas), start=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            
            if (row_idx - 2) % 2 == 0:
                cell.fill = PatternFill(start_color=color_fila_alterna, end_color=color_fila_alterna, fill_type="solid")
    
    for col_idx in range(1, num_columnas + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    ws.freeze_panes = "A2"


def generar_excel_productos_activos():
    productos = productos_repository.listar_productos_activos()

    wb = Workbook()
    ws = wb.active
    ws.title = "Productos Activos"

    columnas = [
        "ID", "Nombre", "Marca", "Categoría", "Precio",
        "Stock", "Creado Por", "Fecha", "Hora",
        "Modificado Por", "Fecha Mod.", "Hora Mod."
    ]
    ws.append(columnas)

    for p in productos:
        fecha, hora = dividir_fecha_hora(p.get("created_at"))
        fecha_mod, hora_mod = dividir_fecha_hora(p.get("modified_at"))

        ws.append([
            p.get("id"),
            p.get("nombre"),
            p.get("marca"),
            p.get("categoria"),
            float(p.get("precio", 0)),
            p.get("stock"),
            p.get("created_by"),
            fecha, hora,
            p.get("modified_by"),
            fecha_mod, hora_mod
        ])
    
    aplicar_estilos_excel(ws, len(columnas))
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=5):
        for cell in row:
            cell.number_format = '$#,##0.00'

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def generar_excel_productos_inactivos():
    productos = productos_repository.listar_productos_inactivos()

    wb = Workbook()
    ws = wb.active
    ws.title = "Productos Inactivos"

    columnas = [
        "ID", "Nombre", "Marca", "Categoría", "Precio",
        "Stock", "Creado Por", "Fecha", "Hora",
        "Modificado Por", "Fecha Mod.", "Hora Mod."
    ]
    ws.append(columnas)

    for p in productos:
        fecha, hora = dividir_fecha_hora(p.get("created_at"))
        fecha_mod, hora_mod = dividir_fecha_hora(p.get("modified_at"))

        ws.append([
            p.get("id"),
            p.get("nombre"),
            p.get("marca"),
            p.get("categoria"),
            float(p.get("precio", 0)),
            p.get("stock"),
            p.get("created_by"),
            fecha, hora,
            p.get("modified_by"),
            fecha_mod, hora_mod
        ])
    
    aplicar_estilos_excel(ws, len(columnas))
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=5):
        for cell in row:
            cell.number_format = '$#,##0.00'

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def generar_excel_movimientos_entrada():
    movimientos = movimientos_repository.listar_entradas_completo()

    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos Entrada"

    columnas = [
        "ID", "Producto", "Precio", "Cantidad", "Monto Total",
        "Tipo", "Motivo", "Creado Por", "Fecha", "Hora"
    ]
    ws.append(columnas)

    for m in movimientos:
        fecha, hora = dividir_fecha_hora(m.get("created_at"))
        producto = m.get("productos") or {}
        precio = producto.get("precio", 0)
        monto_total = float(precio) * int(m.get("cantidad", 0))

        ws.append([
            m.get("id"),
            producto.get("nombre"),
            float(precio),
            m.get("cantidad"),
            monto_total,
            m.get("tipo_movimiento"),
            m.get("motivo"),
            m.get("created_by"),
            fecha, hora
        ])
    
    aplicar_estilos_excel(ws, len(columnas))
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=5):
        for cell in row:
            cell.number_format = '$#,##0.00'

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def generar_excel_movimientos_salida():
    movimientos = movimientos_repository.listar_salidas_completo()

    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos Salida"

    columnas = [
        "ID", "Producto", "Precio", "Cantidad", "Monto Total",
        "Tipo", "Motivo", "Creado Por", "Fecha", "Hora"
    ]
    ws.append(columnas)

    for m in movimientos:
        fecha, hora = dividir_fecha_hora(m.get("created_at"))
        producto = m.get("productos") or {}
        precio = producto.get("precio", 0)
        monto_total = float(precio) * int(m.get("cantidad"))

        ws.append([
            m.get("id"),
            producto.get("nombre"),
            float(precio),
            m.get("cantidad"),
            monto_total,
            m.get("tipo_movimiento"),
            m.get("motivo"),
            m.get("created_by"),
            fecha, hora
        ])
    
    aplicar_estilos_excel(ws, len(columnas))
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=5):
        for cell in row:
            cell.number_format = '$#,##0.00'

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
